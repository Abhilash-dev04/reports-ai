"""Safe synchronization of approved report metadata to the Excel master."""

import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict

from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

EXPECTED_COLUMNS = [
    "Report ID",
    "Job Name",
    "Report Description",
    "Module",
    "Package",
    "Script Name",
    "Output Format",
    "Frequency",
    "Report Type",
    "State",
    "Data Source",
    "Predecessor",
    "Successor",
    "Tables Used",
    "Columns In Tables",
    "QUERY",
    "Filters",
]

FIELD_TO_COLUMN = {
    "report_id": "Report ID",
    "job_name": "Job Name",
    "report_name": "Report Description",
    "functional_area": "Module",
    "package_name": "Package",
    "script_name": "Script Name",
    "output_format": "Output Format",
    "frequency": "Frequency",
    "report_type": "Report Type",
    "state": "State",
    "data_source": "Data Source",
    "predecessor": "Predecessor",
    "successor": "Successor",
    "tables_used": "Tables Used",
    "columns_in_tables": "Columns In Tables",
    "report_query": "QUERY",
    "filters": "Filters",
}


class ExcelSyncError(RuntimeError):
    """Raised when a report cannot be synchronized to Excel."""


def _project_root() -> Path:
    return Path(__file__).resolve().parent.parent.parent


def _resolve_path(environment_name: str, default: str) -> Path:
    configured = os.environ.get(environment_name, default).strip()
    path = Path(configured).expanduser()
    if not path.is_absolute():
        path = _project_root() / path
    return path.resolve()


def get_excel_source_path() -> Path:
    return _resolve_path("EXCEL_SOURCE_PATH", "./data/sample_reports.xlsx")


def get_excel_backup_dir() -> Path:
    return _resolve_path("EXCEL_BACKUP_DIR", "./data/backups")


def _clean(value: object) -> str:
    return str(value or "").strip()


def _safe_error(error: Exception) -> str:
    return f"{type(error).__name__}: {str(error).strip()}"[:2000]


def _validate_headers(worksheet) -> Dict[str, int]:
    headers = {
        _clean(cell.value): index
        for index, cell in enumerate(worksheet[1], start=1)
        if _clean(cell.value)
    }
    missing = [column for column in EXPECTED_COLUMNS if column not in headers]
    if missing:
        raise ExcelSyncError(
            "Excel master is missing required columns: " + ", ".join(missing)
        )
    return headers


def _find_report_rows(worksheet, report_id_column: int, report_id: str) -> list[int]:
    target = report_id.casefold()
    matches = []
    for row_number in range(2, worksheet.max_row + 1):
        current = _clean(worksheet.cell(row=row_number, column=report_id_column).value)
        if current and current.casefold() == target:
            matches.append(row_number)
    return matches


def _create_backup(source_path: Path) -> Path:
    backup_dir = get_excel_backup_dir()
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    backup_path = backup_dir / f"{source_path.stem}_{timestamp}{source_path.suffix}"
    shutil.copy2(source_path, backup_path)
    return backup_path


def sync_report_to_excel(report_data: Dict[str, object]) -> dict:
    """Update or append one report and atomically replace the Excel master."""

    source_path = get_excel_source_path()
    report_id = _clean(report_data.get("report_id"))

    if not report_id:
        raise ExcelSyncError("Report ID is required for Excel synchronization")
    if not source_path.is_file():
        raise ExcelSyncError(f"Excel master was not found: {source_path}")
    if source_path.suffix.lower() != ".xlsx":
        raise ExcelSyncError("EXCEL_SOURCE_PATH must reference an .xlsx workbook")

    temporary_path = source_path.with_name(
        f".{source_path.stem}.sync-{os.getpid()}{source_path.suffix}"
    )

    try:
        workbook = load_workbook(source_path)
        worksheet = workbook.active
        headers = _validate_headers(worksheet)
        matching_rows = _find_report_rows(
            worksheet,
            headers["Report ID"],
            report_id,
        )

        if len(matching_rows) > 1:
            raise ExcelSyncError(
                f"Excel master contains duplicate Report ID '{report_id}'"
            )

        if matching_rows:
            row_number = matching_rows[0]
            action = "updated"
        else:
            row_number = worksheet.max_row + 1
            action = "inserted"

        for field, column_name in FIELD_TO_COLUMN.items():
            worksheet.cell(
                row=row_number,
                column=headers[column_name],
                value=_clean(report_data.get(field)),
            )

        backup_path = _create_backup(source_path)
        workbook.save(temporary_path)
        workbook.close()
        os.replace(temporary_path, source_path)

        return {
            "success": True,
            "action": action,
            "report_id": report_id,
            "row_number": row_number,
            "excel_path": str(source_path),
            "backup_path": str(backup_path),
        }

    except PermissionError as error:
        raise ExcelSyncError(
            "Excel synchronization failed because the workbook is open or locked. "
            "Close sample_reports.xlsx and retry."
        ) from error
    except ExcelSyncError:
        raise
    except Exception as error:
        raise ExcelSyncError(_safe_error(error)) from error
    finally:
        if temporary_path.exists():
            temporary_path.unlink(missing_ok=True)
